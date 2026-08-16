"""
excel_export.py
----------------
Builds a formatted .xlsx workbook (Assumptions, Income Statement,
DCF, Sensitivity) with native Excel charts, from the financials dict
and DCF results produced elsewhere in the app.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
BOLD = Font(bold=True)


def _write_header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws, n_cols):
    for i in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16


def build_workbook(financials: dict, dcf_result: dict, sensitivity: dict, output_path: str):
    wb = Workbook()

    # ---------------- Income Statement sheet ----------------
    ws1 = wb.active
    ws1.title = "Income Statement"
    ws1["A1"] = f"{financials.get('company_name', 'Company')} — Historical Financials"
    ws1["A1"].font = TITLE_FONT
    ws1["A2"] = f"Units: {financials.get('units', '')} {financials.get('currency', '')}"

    years = financials["fiscal_years"]
    _write_header_row(ws1, 4, ["Line item"] + years)

    inc = financials["income_statement"]
    line_items = [
        ("Revenue", inc["revenue"]),
        ("COGS", inc["cogs"]),
        ("Gross Profit", inc["gross_profit"]),
        ("SG&A", inc["sga"]),
        ("EBIT", inc["ebit"]),
        ("Interest Expense", inc["interest_expense"]),
        ("Tax Expense", inc["tax_expense"]),
        ("Net Income", inc["net_income"]),
        ("D&A", inc["depreciation_amortization"]),
    ]
    row = 5
    for label, values in line_items:
        ws1.cell(row=row, column=1, value=label).font = BOLD
        for i, v in enumerate(values):
            ws1.cell(row=row, column=2 + i, value=v)
        row += 1
    _autosize(ws1, len(years) + 1)

    # Revenue chart
    chart = BarChart()
    chart.title = "Revenue by Year"
    chart.y_axis.title = "Revenue"
    data = Reference(ws1, min_col=2, max_col=1 + len(years), min_row=5, max_row=5)
    cats = Reference(ws1, min_col=2, max_col=1 + len(years), min_row=4, max_row=4)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    ws1.add_chart(chart, f"A{row + 2}")

    # ---------------- DCF sheet ----------------
    ws2 = wb.create_sheet("DCF")
    ws2["A1"] = "DCF Valuation (FCFF Approach)"
    ws2["A1"].font = TITLE_FONT

    assumptions = dcf_result["assumptions_used"]
    ws2["A3"] = "Assumptions"
    ws2["A3"].font = BOLD
    assumption_rows = [
        ("EBIT Margin", assumptions["ebit_margin"]),
        ("D&A % of Revenue", assumptions["da_pct_revenue"]),
        ("Capex % of Revenue", assumptions["capex_pct_revenue"]),
        ("Tax Rate", assumptions["tax_rate"]),
        ("WACC", assumptions["wacc"]),
        ("Terminal Growth Rate", assumptions["terminal_growth"]),
    ]
    r = 4
    for label, val in assumption_rows:
        ws2.cell(row=r, column=1, value=label)
        ws2.cell(row=r, column=2, value=val).number_format = "0.0%"
        r += 1

    r += 1
    _write_header_row(
        ws2, r, ["Year", "Revenue", "EBIT", "NOPAT", "D&A", "Capex", "ΔNWC", "FCFF", "PV of FCFF"]
    )
    header_row = r
    r += 1
    forecast_start_row = r
    for row_data in dcf_result["forecast_rows"]:
        ws2.cell(row=r, column=1, value=row_data["year"])
        ws2.cell(row=r, column=2, value=row_data["revenue"])
        ws2.cell(row=r, column=3, value=row_data["ebit"])
        ws2.cell(row=r, column=4, value=row_data["nopat"])
        ws2.cell(row=r, column=5, value=row_data["da"])
        ws2.cell(row=r, column=6, value=row_data["capex"])
        ws2.cell(row=r, column=7, value=row_data["delta_nwc"])
        ws2.cell(row=r, column=8, value=row_data["fcff"])
        ws2.cell(row=r, column=9, value=row_data["pv_fcff"])
        r += 1
    forecast_end_row = r - 1

    r += 1
    summary_rows = [
        ("Sum of PV of FCFF", dcf_result["sum_pv_fcff"]),
        ("Terminal Value", dcf_result["terminal_value"]),
        ("PV of Terminal Value", dcf_result["pv_terminal_value"]),
        ("Enterprise Value", dcf_result["enterprise_value"]),
        ("Net Debt", dcf_result["net_debt"]),
        ("Equity Value", dcf_result["equity_value"]),
        ("Shares Outstanding", dcf_result["shares_outstanding"]),
        ("Implied Share Price", dcf_result["implied_share_price"]),
    ]
    for label, val in summary_rows:
        ws2.cell(row=r, column=1, value=label).font = BOLD
        ws2.cell(row=r, column=2, value=val)
        r += 1

    # FCFF chart
    chart2 = LineChart()
    chart2.title = "Projected FCFF"
    data2 = Reference(ws2, min_col=8, max_col=8, min_row=header_row, max_row=forecast_end_row)
    cats2 = Reference(ws2, min_col=1, max_col=1, min_row=forecast_start_row, max_row=forecast_end_row)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    ws2.add_chart(chart2, f"K4")

    _autosize(ws2, 9)

    # ---------------- Sensitivity sheet ----------------
    ws3 = wb.create_sheet("Sensitivity")
    ws3["A1"] = "Implied Share Price — WACC vs Terminal Growth"
    ws3["A1"].font = TITLE_FONT

    ws3.cell(row=3, column=1, value="WACC \\ g")
    for j, g in enumerate(sensitivity["growth_range"]):
        c = ws3.cell(row=3, column=2 + j, value=g)
        c.number_format = "0.0%"
        c.font = BOLD
    for i, w in enumerate(sensitivity["wacc_range"]):
        c = ws3.cell(row=4 + i, column=1, value=w)
        c.number_format = "0.0%"
        c.font = BOLD
        for j, val in enumerate(sensitivity["grid"][i]):
            ws3.cell(row=4 + i, column=2 + j, value=val)
    _autosize(ws3, len(sensitivity["growth_range"]) + 1)

    wb.save(output_path)
    return output_path
